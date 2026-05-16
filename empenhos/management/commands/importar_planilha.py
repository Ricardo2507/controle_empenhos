from decimal import Decimal
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from empenhos.models import Categoria, Classe, Processo, LancamentoEmpenho

MESES = {
    'Janeiro': 1, 'Fevereiro': 2, 'Março': 3, 'Marco': 3, 'Abril': 4,
    'Maio': 5, 'Junho': 6, 'Julho': 7, 'Agosto': 8, 'Setembro': 9,
    'Outubro': 10, 'Novembro': 11, 'Dezembro': 12,
}


class Command(BaseCommand):
    help = 'Importa dados da planilha de controle de empenhos.'

    def add_arguments(self, parser):
        parser.add_argument('arquivo')

    def handle(self, *args, **options):
        wb = load_workbook(options['arquivo'], data_only=True)
        ws_config = wb['CONFIG']
        ws_proc = wb['PROCESSOS']
        ws_lanc = wb['LANCAMENTOS']

        for row in ws_config.iter_rows(min_row=2, max_row=ws_config.max_row, values_only=True):
            if row[0] and isinstance(row[0], str):
                Categoria.objects.get_or_create(nome=row[0].strip())
            if row[9] and isinstance(row[9], str):
                Categoria.objects.get_or_create(nome=row[9].strip())
            if row[0] and isinstance(row[0], int):
                Classe.objects.get_or_create(codigo=row[0], defaults={'descricao': row[1] or ''})

        for row in ws_proc.iter_rows(min_row=2, values_only=True):
            if row[0]:
                Processo.objects.get_or_create(numero=str(row[0]).strip(), defaults={'observacao': row[1] or ''})

        count = 0
        for row in ws_lanc.iter_rows(min_row=5, values_only=True):
            mes, ano, classe_codigo, categoria_nome, processo_numero, valor_inf, valor_emp, data_lanc, obs, conferido = row[:10]
            if not (mes and ano and classe_codigo and categoria_nome and processo_numero):
                continue
            categoria, _ = Categoria.objects.get_or_create(nome=str(categoria_nome).strip())
            classe, _ = Classe.objects.get_or_create(codigo=int(classe_codigo))
            classe.categorias.add(categoria)
            processo, _ = Processo.objects.get_or_create(numero=str(processo_numero).strip())
            LancamentoEmpenho.objects.create(
                mes=MESES.get(str(mes).strip(), 1),
                ano=int(ano),
                classe=classe,
                categoria=categoria,
                processo=processo,
                valor_informado=Decimal(str(valor_inf or 0)),
                valor_empenhado=Decimal(str(valor_emp or 0)),
                data_lancamento=data_lanc.date() if hasattr(data_lanc, 'date') else data_lanc,
                observacao=obs or '',
                conferido=str(conferido).lower() in ['sim', 'true', '1'],
            )
            count += 1
        self.stdout.write(self.style.SUCCESS(f'Importação concluída: {count} lançamentos importados.'))
